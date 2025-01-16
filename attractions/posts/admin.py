from django.apps import AppConfig
from django.contrib import admin
from django.shortcuts import redirect, render
from django.urls import path
from openpyxl.reader.excel import load_workbook

from .models import (Advice,
                     Category,
                     Comment,
                     Country,
                     Favorite,
                     Post,
                     Tag,
                     Town)

from .forms import XlsxImportForm


class PostConfig(AppConfig):
    default_auto_field = 'django.dbmodels.BigAutoField'
    name = 'post'
    verbose_name = 'Достопримечательности'


class CategoryAdmin(admin.ModelAdmin):
    list_display = ('title', 'description', 'slug')
    change_list_template = 'posts/record_change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-records-from-xlsx/', self.import_records_from_xlsx),
        ]
        return my_urls + urls

    def import_records_from_xlsx(self, request):
        context = admin.site.each_context(request)
        if request.method == 'POST':
            xlsx_file = request.FILES['xlsx_file']

            workbook = load_workbook(filename=xlsx_file, read_only=True)
            worksheet = workbook.active

            records_to_save = []
            for row in worksheet.rows:
                new_obj = self.model(
                    title=row[0].value,
                    description=row[1].value,
                    slug=row[2].value)
                records_to_save.append(new_obj)
            self.model.objects.bulk_create(records_to_save)
            self.message_user(
                request, f'Импортировано строк: {len(records_to_save)}')
            return redirect('admin:posts_category_changelist')
        context['form'] = XlsxImportForm()
        return render(request, 'posts/add_records_form.html', context=context)


class CountryAdmin(admin.ModelAdmin):
    list_display = ('title', 'description', 'slug')
    change_list_template = 'posts/record_change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-records-from-xlsx/', self.import_records_from_xlsx),
        ]
        return my_urls + urls

    def import_records_from_xlsx(self, request):
        context = admin.site.each_context(request)
        if request.method == 'POST':
            xlsx_file = request.FILES['xlsx_file']

            workbook = load_workbook(filename=xlsx_file, read_only=True)
            worksheet = workbook.active

            records_to_save = []
            for row in worksheet.rows:
                new_obj = self.model(
                    title=row[0].value,
                    description=row[1].value,
                    slug=row[2].value)
                records_to_save.append(new_obj)
            self.model.objects.bulk_create(records_to_save)
            self.message_user(
                request, f'Импортировано строк: {len(records_to_save)}')
            return redirect('admin:posts_country_changelist')
        context['form'] = XlsxImportForm()
        return render(request, 'posts/add_records_form.html', context=context)


class TownAdmin(admin.ModelAdmin):
    list_display = ('title', 'description', 'slug', 'country')
    change_list_template = 'posts/record_change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-records-from-xlsx/', self.import_records_from_xlsx),
        ]
        return my_urls + urls

    def import_records_from_xlsx(self, request):
        context = admin.site.each_context(request)
        if request.method == 'POST':
            xlsx_file = request.FILES['xlsx_file']

            workbook = load_workbook(filename=xlsx_file, read_only=True)
            worksheet = workbook.active

            records_to_save = []
            for row in worksheet.rows:
                new_obj = self.model(
                    title=row[0].value,
                    description=row[1].value,
                    slug=row[2].value,
                    country=Country.objects.get(id=row[3].value))
                records_to_save.append(new_obj)
            self.model.objects.bulk_create(records_to_save)
            self.message_user(
                request, f'Импортировано строк: {len(records_to_save)}')
            return redirect('admin:posts_town_changelist')
        context['form'] = XlsxImportForm()
        return render(request, 'posts/add_records_form.html', context=context)


admin.site.register(Category, CategoryAdmin)
admin.site.register(Advice)
admin.site.register(Comment)
admin.site.register(Country, CountryAdmin)
admin.site.register(Favorite)
admin.site.register(Post)
admin.site.register(Tag)
admin.site.register(Town, TownAdmin)
